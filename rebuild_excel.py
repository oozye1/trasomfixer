#!/usr/bin/env python3
"""Rebuild TitanPour_Transom_Calculator.xlsx with pentagon shape formulas.

Matches TransomCalculator.jsx calcs exactly.

Sheets:
  1. Area & Volume — main calculator (pentagon shape, no weight)
  2. Rod Grid — individual rod lengths (H rods vary in tri zone, V rods vary by position)
  3. Mix Calculator — batch mix ratios
  4. Reference — formula documentation
"""
import sys, io, math
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
hdr_font = Font(bold=True, size=12, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1E3A5F")
label_font = Font(size=11)
val_font = Font(bold=True, size=11, color="1E3A5F")
unit_font = Font(size=9, color="888888")
section_font = Font(bold=True, size=11, color="1E3A5F")
section_fill = PatternFill("solid", fgColor="E8F0FE")
note_font = Font(size=9, color="666666", italic=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def style_header(ws, row, text, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')

def add_row(ws, row, label, formula_or_val, unit="", note=""):
    ws.cell(row=row, column=1, value=label).font = label_font
    c = ws.cell(row=row, column=2, value=formula_or_val)
    c.font = val_font; c.alignment = Alignment(horizontal='right')
    if unit:
        ws.cell(row=row, column=3, value=unit).font = unit_font
    if note:
        ws.cell(row=row, column=5, value=note).font = note_font

def section_row(ws, row, text, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = section_font; c.fill = section_fill

# =============================================================================
# SHEET 1: Area & Volume
# =============================================================================
ws1 = wb.active
ws1.title = "Area & Volume"
ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 16
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 4
ws1.column_dimensions['E'].width = 40

style_header(ws1, 1, "TitanPour Transom Calculator - Area & Volume")

# ── Inputs ──
section_row(ws1, 3, "INPUTS")
add_row(ws1, 4, "Transom Width", 1800, "mm")
add_row(ws1, 5, "Centre Height (vertical)", 508, "mm", "Tallest point (middle)")
add_row(ws1, 6, "Side Height (vertical)", 350, "mm", "Shortest point (edges)")
add_row(ws1, 7, "Rake Angle", 20, "deg")
add_row(ws1, 8, "Thickness", 50, "mm")
add_row(ws1, 9, "Shell Thickness", 6, "mm", "Existing outer GRP skin")
add_row(ws1, 10, "Wastage", 10, "%")
add_row(ws1, 11, "Cutout Width", 660, "mm")
add_row(ws1, 12, "Cutout Height", 380, "mm")
add_row(ws1, 13, "Cutout Count", 1, "")
add_row(ws1, 14, "Material Density", 1550, "kg/m3", "TitanPour default")

# ── Slope Heights ──
section_row(ws1, 16, "SLOPE HEIGHTS")
add_row(ws1, 17, "Centre Slope Height", '=B5/COS(RADIANS(B7))', "mm", "= centreHeight / cos(angle)")
add_row(ws1, 18, "Side Slope Height", '=B6/COS(RADIANS(B7))', "mm", "= sideHeight / cos(angle)")

# ── Area ──
section_row(ws1, 20, "AREA (Pentagon / Trapezoid)")
add_row(ws1, 21, "Gross Area", '=B4*(B17+B18)/2/1000000', "m2", "= W x (centreSlopeH + sideSlopeH) / 2")
add_row(ws1, 22, "Cutout Slope Height", '=IF(B13>0,B12/COS(RADIANS(B7)),0)', "mm")
add_row(ws1, 23, "Single Cutout Area", '=IF(B13>0,B11*B22/1000000,0)', "m2")
add_row(ws1, 24, "Total Cutout Area", '=B23*B13', "m2")
add_row(ws1, 25, "Net Area", '=B21-B24', "m2")

# ── Volume ──
section_row(ws1, 27, "VOLUME")
add_row(ws1, 28, "Cavity Volume", '=B25*B8', "litres", "= netArea(m2) x thickness(mm) = litres")
add_row(ws1, 29, "Rod Displacement", "='Rod Grid'!B33", "litres", "From Rod Grid sheet")
add_row(ws1, 30, "Pour Volume (net)", '=B28-B29', "litres", "Cavity minus rods")
add_row(ws1, 31, "Resin (no wastage)", '=B30', "litres")
add_row(ws1, 32, "Resin (with wastage)", '=B31*(1+B10/100)', "litres")
add_row(ws1, 33, "Order Weight", '=B32*B14/1000', "kg", "= litres x density")

# ── Pour Depth ──
section_row(ws1, 35, "POUR DEPTH")
add_row(ws1, 36, "Pour Depth", '=B8-B9', "mm", "= thickness - shell")

print("Sheet 1 (Area & Volume): done")

# =============================================================================
# SHEET 2: Rod Grid
# =============================================================================
ws2 = wb.create_sheet("Rod Grid")
ws2.column_dimensions['A'].width = 24
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 4
ws2.column_dimensions['E'].width = 36
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 8

style_header(ws2, 1, "Rod Grid Calculation", 7)

# ── Rod Inputs ──
section_row(ws2, 3, "ROD INPUTS")
add_row(ws2, 4, "Rod Spacing", 70, "mm")
add_row(ws2, 5, "H Rod Diameter", 7, "mm")
add_row(ws2, 6, "V Rod Diameter", 7, "mm")
add_row(ws2, 7, "Rod Gap (H rods)", 30, "mm", "Clearance each side")
add_row(ws2, 8, "Min Cover", 10, "mm")

# ── Counts (use linked values from Area sheet) ──
section_row(ws2, 10, "ROD COUNTS")
# Reference centre/side slope height from Sheet 1
add_row(ws2, 11, "Centre Slope Height", "='Area & Volume'!B17", "mm")
add_row(ws2, 12, "Side Slope Height", "='Area & Volume'!B18", "mm")
add_row(ws2, 13, "Width", "='Area & Volume'!B4", "mm")
add_row(ws2, 14, "Centre Height (vert)", "='Area & Volume'!B5", "mm")
add_row(ws2, 15, "Side Height (vert)", "='Area & Volume'!B6", "mm")
add_row(ws2, 16, "Cos(Angle)", '=COS(RADIANS(\'Area & Volume\'!B7))', "")
add_row(ws2, 17, "H Rod Count", '=FLOOR(B11/B4,1)+1', "", "Based on centre slope height")
add_row(ws2, 18, "V Rod Count", '=FLOOR(B13/B4,1)+1', "", "Based on width")

# ── H Rod Cut List (individual rods) ──
section_row(ws2, 20, "H ROD CUT LIST (individual lengths)", 7)
ws2.cell(row=21, column=1, value="Rod #").font = Font(bold=True, size=10)
ws2.cell(row=21, column=2, value="Slope Y (mm)").font = Font(bold=True, size=10)
ws2.cell(row=21, column=3, value="Vert Y (mm)").font = Font(bold=True, size=10)
ws2.cell(row=21, column=4, value="Zone").font = Font(bold=True, size=10)
ws2.cell(row=21, column=5, value="Panel Width (mm)").font = Font(bold=True, size=10)
ws2.cell(row=21, column=6, value="Rod Length (mm)").font = Font(bold=True, size=10)

# Generate up to 20 H rod rows (more than enough for typical transoms)
MAX_H_RODS = 20
for i in range(MAX_H_RODS):
    r = 22 + i
    rod_num = i + 1
    # Only show if rod_num <= hRodCount
    ws2.cell(row=r, column=1, value=f'=IF({rod_num}<=B17,{rod_num},"")')
    # Slope Y = (rod_num - 1) * spacing
    ws2.cell(row=r, column=2, value=f'=IF({rod_num}<=B17,({rod_num}-1)*B4,"")')
    # Vert Y = slopeY * cos(angle)
    ws2.cell(row=r, column=3, value=f'=IF({rod_num}<=B17,B{r}*B16,"")')
    # Zone: RECT if vertY <= sideHeight, else TRI
    ws2.cell(row=r, column=4, value=f'=IF({rod_num}<=B17,IF(C{r}<=B15,"RECT","TRI"),"")')
    # Panel width: full width in RECT zone, narrowing in TRI zone
    ws2.cell(row=r, column=5, value=f'=IF({rod_num}<=B17,IF(C{r}<=B15,B13,B13*(B14-C{r})/(B14-B15)),"")')
    # Rod length: panel width - gap, min 0
    ws2.cell(row=r, column=6, value=f'=IF({rod_num}<=B17,MAX(E{r}-B7,0),"")')

# ── V Rod Cut List ──
v_start = 22 + MAX_H_RODS + 2  # = row 44
section_row(ws2, v_start, "V ROD CUT LIST (individual lengths)", 7)
v_hdr = v_start + 1
ws2.cell(row=v_hdr, column=1, value="Rod #").font = Font(bold=True, size=10)
ws2.cell(row=v_hdr, column=2, value="X pos (mm)").font = Font(bold=True, size=10)
ws2.cell(row=v_hdr, column=3, value="t (0-1)").font = Font(bold=True, size=10)
ws2.cell(row=v_hdr, column=4, value="").font = Font(bold=True, size=10)
ws2.cell(row=v_hdr, column=5, value="Vert Height (mm)").font = Font(bold=True, size=10)
ws2.cell(row=v_hdr, column=6, value="Slope Length (mm)").font = Font(bold=True, size=10)

MAX_V_RODS = 30
for i in range(MAX_V_RODS):
    r = v_hdr + 1 + i
    rod_num = i + 1
    # Only show if rod_num <= vRodCount
    ws2.cell(row=r, column=1, value=f'=IF({rod_num}<=B18,{rod_num},"")')
    # X position = (rod_num - 1) * spacing
    ws2.cell(row=r, column=2, value=f'=IF({rod_num}<=B18,({rod_num}-1)*B4,"")')
    # t = 1 - ABS(2*x/width - 1); 0 at edges, 1 at centre
    ws2.cell(row=r, column=3, value=f'=IF({rod_num}<=B18,1-ABS(2*B{r}/B13-1),"")')
    # Vert height = sideH + (centreH - sideH) * t
    ws2.cell(row=r, column=5, value=f'=IF({rod_num}<=B18,B15+(B14-B15)*C{r},"")')
    # Slope length = height / cos(angle)
    ws2.cell(row=r, column=6, value=f'=IF({rod_num}<=B18,E{r}/B16,"")')

# ── Totals ──
totals_row = v_hdr + 1 + MAX_V_RODS + 1  # after V rods
section_row(ws2, totals_row, "ROD TOTALS", 7)

h_data_start = 22
h_data_end = 22 + MAX_H_RODS - 1
v_data_start = v_hdr + 1
v_data_end = v_hdr + MAX_V_RODS

add_row(ws2, totals_row + 1, "Total H Rod Length", f'=SUMPRODUCT((F{h_data_start}:F{h_data_end}<>"")*F{h_data_start}:F{h_data_end})/1000', "m")
add_row(ws2, totals_row + 2, "Total V Rod Length", f'=SUMPRODUCT((F{v_data_start}:F{v_data_end}<>"")*F{v_data_start}:F{v_data_end})/1000', "m")
add_row(ws2, totals_row + 3, "Total Rod Length", f'=B{totals_row+1}+B{totals_row+2}', "m")
add_row(ws2, totals_row + 4, "Total Rod Count", '=B17+B18', "rods")

# Rod displacement
add_row(ws2, totals_row + 6, "H Rod Cross Section", '=PI()*(B5/2)^2', "mm2")
add_row(ws2, totals_row + 7, "V Rod Cross Section", '=PI()*(B6/2)^2', "mm2")
add_row(ws2, totals_row + 8, "H Rod Volume", f'=B{totals_row+6}*B{totals_row+1}*1000', "mm3", "cross section x length(m) x 1000")
add_row(ws2, totals_row + 9, "V Rod Volume", f'=B{totals_row+7}*B{totals_row+2}*1000', "mm3")
add_row(ws2, totals_row + 10, "Total Rod Displacement", f'=(B{totals_row+8}+B{totals_row+9})/1000000', "litres")

# Store the displacement row number for the cross-sheet reference
disp_row = totals_row + 10
# The Area & Volume sheet references 'Rod Grid'!B33 — let's also put the value there
# Actually, let's check: totals_row = v_hdr + 1 + 30 + 1 = 46 + 31 + 1 = ... let me compute
# v_start = 44, v_hdr = 45, v_data rows = 46..75, totals_row = 77
# disp_row = 77 + 10 = 87
# That's way beyond B33. We need to put the displacement at B33 for the cross-sheet ref.
# Let me restructure to put totals closer.

# Actually, let's just add a named reference at B33
# Wait - the Area & Volume sheet references 'Rod Grid'!B33 for rod displacement
# Let's put it there explicitly

print(f"  H rod data: rows {h_data_start}-{h_data_end}")
print(f"  V rod data: rows {v_data_start}-{v_data_end}")
print(f"  Totals start: row {totals_row}")
print(f"  Displacement at: row {disp_row}")

# We need to adjust. Let me put the displacement value at B33 as well
ws2.cell(row=33, column=1, value="Rod Displacement (linked)").font = label_font
ws2.cell(row=33, column=2, value=f'=B{disp_row}').font = val_font
ws2.cell(row=33, column=3, value="litres").font = unit_font

# ── Cover Analysis ──
cover_row = totals_row + 12
section_row(ws2, cover_row, "COVER ANALYSIS", 7)
add_row(ws2, cover_row + 1, "Pour Depth", "='Area & Volume'!B36", "mm")
add_row(ws2, cover_row + 2, "Stacked Height", '=B5+B6', "mm", "H diam + V diam")
add_row(ws2, cover_row + 3, "Stack Cover", f'=(B{cover_row+1}-B{cover_row+2})/2', "mm", "(pourDepth - stacked) / 2")
add_row(ws2, cover_row + 4, "Stack Fits?", f'=IF(B{cover_row+3}>=B8,"YES","NO")', "", ">= min cover")
add_row(ws2, cover_row + 5, "Min Thickness for Stack", f'=B{cover_row+2}+B8*2+\'Area & Volume\'!B9', "mm")

print("Sheet 2 (Rod Grid): done")

# =============================================================================
# SHEET 3: Mix Calculator
# =============================================================================
ws3 = wb.create_sheet("Mix Calculator")
ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 4
ws3.column_dimensions['E'].width = 36

style_header(ws3, 1, "Mix Calculator")

section_row(ws3, 3, "MIX INPUTS")
add_row(ws3, 4, "Batch Weight", 1000, "g")
add_row(ws3, 5, "Catalyst %", 1.5, "% of resin")

section_row(ws3, 7, "MIX FRACTIONS")
add_row(ws3, 8, "Resin Fraction", 0.58, "", "58%")
add_row(ws3, 9, "Talc Fraction", 0.30, "", "30%")
add_row(ws3, 10, "Glass Fraction", 0.10, "", "10%")
add_row(ws3, 11, "Catalyst+Retarder Frac", 0.02, "", "2%")

section_row(ws3, 13, "MIX WEIGHTS")
add_row(ws3, 14, "Resin Mass", '=B4*B8', "g")
add_row(ws3, 15, "Talc Mass", '=B4*B9', "g")
add_row(ws3, 16, "Glass Mass", '=B4*B10', "g")
add_row(ws3, 17, "Catalyst Mass", '=B14*B5/100', "g", "= resinMass x catalyst%")
add_row(ws3, 18, "Retarder Mass", '=B14*0.00025', "g", "0.025% of resin")
add_row(ws3, 19, "Retarder (grams)", '=B18', "g")

section_row(ws3, 21, "CATALYST OPTIONS")
add_row(ws3, 22, "1.2% (hot >26C)", '=B14*0.012', "g")
add_row(ws3, 23, "1.5% (normal 18-26C)", '=B14*0.015', "g")
add_row(ws3, 24, "2.0% (cold <18C)", '=B14*0.020', "g")

section_row(ws3, 26, "POUR REQUIREMENTS")
add_row(ws3, 27, "Cavity Volume", "='Area & Volume'!B28", "litres")
add_row(ws3, 28, "Rod Displacement", "='Area & Volume'!B29", "litres")
add_row(ws3, 29, "Pour Volume Needed", "='Area & Volume'!B30", "litres")
add_row(ws3, 30, "Pour Weight", "=B29*'Area & Volume'!B14/1000", "kg", "= litres x density")

print("Sheet 3 (Mix Calculator): done")

# =============================================================================
# SHEET 4: Reference
# =============================================================================
ws4 = wb.create_sheet("Reference")
ws4.column_dimensions['A'].width = 36
ws4.column_dimensions['B'].width = 60

style_header(ws4, 1, "Calculation Formulas Reference", 2)

formulas = [
    ("Slope Height", "Vertical Height / cos(Rake Angle)"),
    ("Gross Area (pentagon)", "Width x (Centre Slope H + Side Slope H) / 2"),
    ("Net Area", "Gross Area - Cutout Area"),
    ("Cavity Volume", "Net Area (m2) x Thickness (mm) = litres"),
    ("Pour Volume", "Cavity Volume - Rod Displacement"),
    ("", ""),
    ("H Rod Count", "floor(Centre Slope Height / Spacing) + 1"),
    ("H Rod Length (rect zone)", "Width - 30mm  (where vertY <= sideHeight)"),
    ("H Rod Length (tri zone)", "Width x (centreH - vertY)/(centreH - sideH) - 30mm"),
    ("", ""),
    ("V Rod Count", "floor(Width / Spacing) + 1"),
    ("V Rod Height at x", "sideH + (centreH - sideH) x (1 - |2x/W - 1|)"),
    ("V Rod Slope Length", "Height at x / cos(Rake Angle)"),
    ("", ""),
    ("Rod Displacement", "PI x (d/2)^2 x totalLength"),
    ("Pour Depth", "Thickness - Shell Thickness"),
    ("Stack Cover", "(Pour Depth - H diam - V diam) / 2"),
    ("Min Thickness", "H diam + V diam + 2 x min cover + shell"),
    ("", ""),
    ("Mix: Resin Mass", "Batch Weight x 0.58"),
    ("Mix: Talc Mass", "Batch Weight x 0.30"),
    ("Mix: Glass Mass", "Batch Weight x 0.10"),
    ("Mix: Catalyst", "Resin Mass x catalyst% (applies to RESIN only)"),
    ("Mix: Retarder", "Resin Mass x 0.025%"),
    ("", ""),
    ("Unit: 1 m2 x 1 mm", "1 litre"),
    ("Rod sizing rule of thumb", "(Thickness - Shell) / 3 / 2 = rod diameter per rod"),
]

for i, (label, formula) in enumerate(formulas):
    r = 3 + i
    ws4.cell(row=r, column=1, value=label).font = Font(bold=True if label else False, size=10)
    ws4.cell(row=r, column=2, value=formula).font = Font(size=10, color="444444")

print("Sheet 4 (Reference): done")

# =============================================================================
# SAVE
# =============================================================================
path = r"f:\transom designer\TitanPour_Transom_Calculator.xlsx"
wb.save(path)
print(f"\nSaved: {path}")

# =============================================================================
# VERIFY: read back and check formulas
# =============================================================================
from openpyxl import load_workbook
wb2 = load_workbook(path)

print("\n=== VERIFICATION ===")
errors = 0
for name in wb2.sheetnames:
    ws = wb2[name]
    formula_count = 0
    text_formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula_count += 1
            elif cell.data_type == 's' and isinstance(cell.value, str) and cell.value.startswith('='):
                text_formula_count += 1
                print(f"  ERROR: {name}!{cell.coordinate} stored as TEXT: {cell.value}")
                errors += 1
    print(f"  {name}: {formula_count} formulas, {text_formula_count} text-formulas")

if errors == 0:
    print("\nAll formulas stored correctly (no text-formulas found)")
else:
    print(f"\n{errors} ERRORS found - formulas stored as text!")

# Count total formulas
total = sum(1 for name in wb2.sheetnames for row in wb2[name].iter_rows() for c in row if c.data_type == 'f')
print(f"Total formulas: {total}")
